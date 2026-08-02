from collections import Counter, defaultdict
from pathlib import Path
import sqlite3

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = Path(r"D:\Mywechatlog\xwechat_files\wxid_06l166che9wp22_76f7\msg\file\2026-08")
SOURCE_PATH = sorted(SOURCE_DIR.glob("*_v2(2).xlsx"))[0]
DB_PATH = ROOT / "data" / "real-catalog-campus-2026-07-27-v2.sqlite"
OUTPUT_PATH = ROOT / "artifacts" / "\u83dc\u54c1\u5206\u7c7b\u6807\u51c6_\u6700\u7ec8\u7248_2026-08.xlsx"

TYPE_LABELS = {
    "meal": "\u9910\u98df",
    "snack": "\u5c0f\u5403",
    "beverage": "\u996e\u54c1",
    "addon": "\u52a0\u8d2d",
    "variant": "\u89c4\u683c/\u53e3\u5473\u9009\u9879",
    "fee": "\u8d39\u7528\u9879",
    "section": "\u76ee\u5f55\u5206\u7ec4",
}
PUBLIC_TYPES = {"meal", "snack", "beverage"}

FINAL_CATEGORIES = {
    "meal": [
        "\u9762\u98df/\u7c89\u7c7b", "\u7c73\u996d/\u76d6\u996d", "\u65e9\u9910\u9762\u70b9", "\u997a\u5b50/\u9984\u9968", "\u7ca5\u54c1",
        "\u5bb6\u5e38\u7092\u83dc", "\u6c64\u7fb9", "\u706b\u9505/\u9ebb\u8fa3\u70eb", "\u6c34\u716e/\u5e72\u9505", "\u7802\u9505/\u7172\u7c7b",
        "\u70e4\u9c7c", "\u6c49\u5821/\u70b8\u9e21", "\u70e7\u70e4/\u70b8\u7269", "\u7ec4\u5408\u5957\u9910", "\u6c99\u62c9/\u8f7b\u98df", "\u84b8\u83dc", "\u751c\u54c1", "\u5176\u4ed6\u9910\u98df",
    ],
    "snack": ["\u70e7\u70e4/\u70b8\u7269", "\u6c49\u5821/\u70b8\u9e21", "\u5364\u5473/\u9e2d\u8d27", "\u751c\u54c1", "\u5c0f\u5403"],
    "beverage": ["\u5976\u8336", "\u996e\u54c1"],
    "addon": ["\u4e3b\u98df\u52a0\u8d2d", "\u8089\u7c7b\u52a0\u8d2d", "\u852c\u83dc/\u8c46\u5236\u54c1\u52a0\u8d2d", "\u706b\u9505\u914d\u83dc", "\u5176\u4ed6\u52a0\u8d2d"],
    "variant": ["\u89c4\u683c/\u53e3\u5473\u9009\u9879"],
    "fee": ["\u9910\u5177\u8d39"],
    "section": ["\u76ee\u5f55\u5206\u7ec4"],
}


def canonical_category(item_type, name, source_category, v2_category):
    name = str(name or "")
    source_category = str(source_category or "")
    v2_category = str(v2_category or "")
    category = v2_category or source_category

    if item_type == "section":
        return "\u76ee\u5f55\u5206\u7ec4"
    if item_type == "fee":
        return "\u9910\u5177\u8d39"
    if item_type == "variant":
        return "\u89c4\u683c/\u53e3\u5473\u9009\u9879"
    if item_type == "beverage":
        return "\u5976\u8336" if category == "\u5976\u8336" else "\u996e\u54c1"
    if item_type == "addon":
        if category in {"\u9762\u98df\u52a0\u8d2d", "\u4e3b\u98df\u52a0\u8d2d"}:
            return "\u4e3b\u98df\u52a0\u8d2d"
        if category in {"\u8089\u7c7b\u52a0\u8d2d", "\u852c\u83dc/\u8c46\u5236\u54c1\u52a0\u8d2d", "\u706b\u9505\u914d\u83dc"}:
            return category
        return "\u5176\u4ed6\u52a0\u8d2d"
    if item_type == "snack":
        if category == "\u6c49\u5821" or "\u6c49\u5821" in name:
            return "\u6c49\u5821/\u70b8\u9e21"
        if category in {"\u9762\u98df/\u7c89\u7c7b", "\u997c\u7c7b", "\u5c0f\u5403"}:
            return "\u5c0f\u5403"
        if category in {"\u70e7\u70e4/\u70b8\u7269", "\u5364\u5473/\u9e2d\u8d27", "\u751c\u54c1"}:
            return category
        return "\u5c0f\u5403"

    if any(token in name for token in ("\u70e4\u9c7c", "\u70e4\u8349\u9c7c", "\u70e4\u91cc\u9c7c", "\u6e05\u6c5f\u9c7c")):
        return "\u70e4\u9c7c"
    if category in {"\u997c\u7c7b", "\u65e9\u9910\u9762\u70b9"}:
        return "\u65e9\u9910\u9762\u70b9"
    mapping = {
        "\u9762\u98df/\u7c89\u7c7b": "\u9762\u98df/\u7c89\u7c7b", "\u7c73\u996d/\u76d6\u996d": "\u7c73\u996d/\u76d6\u996d", "\u997a\u5b50/\u9984\u9968": "\u997a\u5b50/\u9984\u9968",
        "\u5bb6\u5e38\u7092\u83dc": "\u5bb6\u5e38\u7092\u83dc", "\u6c64\u7fb9": "\u6c64\u7fb9", "\u706b\u9505/\u9ebb\u8fa3\u70eb": "\u706b\u9505/\u9ebb\u8fa3\u70eb",
        "\u6c34\u716e/\u5e72\u9505": "\u6c34\u716e/\u5e72\u9505", "\u7802\u9505/\u7172\u7c7b": "\u7802\u9505/\u7172\u7c7b",
        "\u70e4\u9c7c": "\u70e4\u9c7c", "\u6c49\u5821": "\u6c49\u5821/\u70b8\u9e21", "\u70e7\u70e4/\u70b8\u7269": "\u70e7\u70e4/\u70b8\u7269",
        "\u7ec4\u5408\u5957\u9910": "\u7ec4\u5408\u5957\u9910", "\u6c99\u62c9/\u8f7b\u98df": "\u6c99\u62c9/\u8f7b\u98df", "\u84b8\u83dc": "\u84b8\u83dc",
        "\u7ca5\u54c1": "\u7ca5\u54c1", "\u751c\u54c1": "\u751c\u54c1", "\u5bb6\u5e38\u70ed\u83dc": "\u5bb6\u5e38\u7092\u83dc",
        "\u5e72\u9505\u83dc": "\u6c34\u716e/\u5e72\u9505", "\u6c34\u716e\u83dc": "\u6c34\u716e/\u5e72\u9505", "\u7802\u9505\u7172\u7c7b": "\u7802\u9505/\u7172\u7c7b",
        "\u706b\u9505\u9ebb\u8fa3\u70eb": "\u706b\u9505/\u9ebb\u8fa3\u70eb", "\u7ec4\u5408\u5957\u9910": "\u7ec4\u5408\u5957\u9910", "\u8f7b\u98df\u7b80\u9910": "\u6c99\u62c9/\u8f7b\u98df",
    }
    return mapping.get(category, "\u5176\u4ed6\u9910\u98df")


def safe(value):
    return "" if value is None else value


def main():
    wb_source = load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    first = list(wb_source.worksheets[0].iter_rows(values_only=True))
    second = list(wb_source.worksheets[1].iter_rows(values_only=True))
    third = list(wb_source.worksheets[2].iter_rows(values_only=True))

    first_by_id = {r[0]: r for r in first[1:] if isinstance(r[0], str) and r[0].startswith("dish-")}
    second_by_id = {r[0]: r for r in second[1:] if isinstance(r[0], str) and r[0].startswith("dish-")}
    third_by_id = {r[0]: r for r in third[1:] if isinstance(r[0], str) and r[0].startswith("dish-")}

    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    db_rows = db.execute(
        """
        SELECT d.id, d.name, d.price, d.status, d.tenant_id, d.stall_id,
               s.name AS stall_name, s.floor, s.parent_id,
               c.name AS canteen_name, p.name AS parent_name
        FROM dishes d
        LEFT JOIN stalls s ON s.id = d.stall_id
        LEFT JOIN stalls p ON p.id = s.parent_id
        LEFT JOIN canteens c ON c.id = s.canteen_id
        ORDER BY d.id
        """
    ).fetchall()
    db_by_id = {r["id"]: r for r in db_rows}

    records = []
    for dish_id, row in first_by_id.items():
        item_type = str(row[6])
        v2 = second_by_id.get(dish_id)
        period = third_by_id.get(dish_id)
        original_category = safe(row[7])
        v2_category = safe(v2[8]) if v2 else original_category
        category = canonical_category(item_type, row[1], original_category, v2_category)
        records.append({
            "id": dish_id, "name": safe(row[1]), "canteen": safe(row[2]), "restaurant": safe(row[3]), "stall": safe(row[4]),
            "price": safe(row[5]), "item_type": item_type, "original_category": original_category,
            "category": category, "period": safe(period[8]) if period else "\u5f85\u6838\u9a8c",
            "source": "v2\u5206\u7c7b\u8868", "retrieval": "\u662f" if item_type in PUBLIC_TYPES else "\u5426",
            "review": "\u5f85\u6309\u6570\u636e\u5e93\u5ba1\u6838\u72b6\u6001\u786e\u8ba4" if item_type in PUBLIC_TYPES else "\u7ed3\u6784\u9879\uff0c\u6392\u9664\u5b66\u751f\u68c0\u7d22",
        })

    excel_ids = set(first_by_id)
    for row in db_rows:
        if row["id"] in excel_ids:
            continue
        records.append({
            "id": row["id"], "name": row["name"], "canteen": row["canteen_name"] or "", "restaurant": row["parent_name"] or "",
            "stall": row["stall_name"] or "", "price": row["price"] or "", "item_type": "section", "original_category": "\u76ee\u5f55\u5206\u7ec4",
            "category": "\u76ee\u5f55\u5206\u7ec4", "period": "\u4e0d\u9002\u7528", "source": "\u771f\u5b9e\u76ee\u5f55\u5feb\u7167\uff08v2\uff09\u8865\u56de",
            "retrieval": "\u5426", "review": "\u7ed3\u6784\u5206\u7ec4\uff0c\u4e0d\u4f5c\u72ec\u7acb\u83dc\u54c1",
        })

    records.sort(key=lambda r: r["id"])
    assert len(records) == 2563, len(records)
    assert len({r["id"] for r in records}) == 2563

    wb = Workbook()
    standard = wb.active
    standard.title = "\u5206\u7c7b\u6807\u51c6"
    data_ws = wb.create_sheet("\u6700\u7ec8\u5206\u7c7b\u6570\u636e")
    review_ws = wb.create_sheet("\u5f85\u4eba\u5de5\u6838\u9a8c")
    summary_ws = wb.create_sheet("\u7edf\u8ba1\u6458\u8981")
    handoff_ws = wb.create_sheet("\u4ea4\u63a5\u89c4\u5219")

    thin = Side(style="thin", color="D9E2F3")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(color="FFFFFF", bold=True)

    def style_sheet(ws, widths=None):
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(bottom=thin)
        if widths:
            for col, width in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 32

    standard_headers = ["\u7ef4\u5ea6", "\u4ee3\u7801", "\u4e2d\u6587\u540d\u79f0", "\u7528\u9014", "\u662f\u5426\u5b66\u751f\u7aef\u72ec\u7acb\u68c0\u7d22", "\u5206\u7c7b\u89c4\u5219"]
    standard.append(standard_headers)
    standard_rows = [
        ["\u4f4d\u7f6e\u5c42\u7ea7", "location", "\u98df\u5802 \u2192 \u9910\u5385/\u697c\u5c42 \u2192 \u6863\u53e3 \u2192 \u83dc\u54c1", "\u5b9a\u4f4d\u548c\u5bfc\u822a", "\u662f", "\u4e0e\u83dc\u54c1\u7c7b\u578b\u5e76\u5217\uff0c\u4e0d\u6539\u53d8\u7c7b\u578b"],
        ["\u4e1a\u52a1\u5206\u533a", "item_type", "\u9910\u98df / \u5c0f\u5403 / \u996e\u54c1 / \u52a0\u8d2d / \u89c4\u683c / \u8d39\u7528 / \u76ee\u5f55\u5206\u7ec4", "\u51b3\u5b9a\u641c\u7d22\u5165\u53e3", "\u4ec5\u9910\u98df\u3001\u5c0f\u5403\u3001\u996e\u54c1", "\u52a0\u8d2d\u3001\u8d39\u7528\u3001\u89c4\u683c\u548c\u5206\u7ec4\u4e0d\u4f5c\u72ec\u7acb\u83dc\u54c1"],
        ["\u83dc\u54c1\u5206\u7c7b", "category", "\u5404\u4e1a\u52a1\u5206\u533a\u7684\u4e8c\u7ea7\u7c7b\u522b", "\u5206\u7c7b\u7b5b\u9009\u548c\u5c55\u793a", "\u662f", "\u53ea\u5728\u5f53\u524d\u4e00\u7ea7\u5206\u533a\u5185\u7b5b\u9009"],
        ["\u9910\u6b21\u5c5e\u6027", "meal_period", "\u65e9\u9910 / \u5348\u9910\u665a\u9910 / \u5168\u5929 / \u5f85\u6838\u9a8c", "\u8f85\u52a9\u7b5b\u9009\u548c\u6807\u7b7e", "\u5426", "\u539f\u59cb\u8d44\u6599\u672a\u8bc1\u660e\u4f9b\u5e94\u65f6\u6bb5\uff0c\u4e0d\u5f53\u4f5c\u4eca\u65e5\u4f9b\u5e94\u4e8b\u5b9e"],
    ]
    for row in standard_rows:
        standard.append(row)
    standard.append([])
    standard.append(["\u4e00\u7ea7\u4e1a\u52a1\u5206\u533a", "\u5bf9\u5e94\u4e2d\u6587", "\u68c0\u7d22\u653f\u7b56", "\u8bf4\u660e", "", ""])
    for code, label in TYPE_LABELS.items():
        policy = "\u9ed8\u8ba4\u5165\u53e3" if code == "meal" else ("\u660e\u786e\u5207\u6362\u540e\u68c0\u7d22" if code in {"snack", "beverage"} else "\u4e0d\u5165\u5b66\u751f\u7aef\u72ec\u7acb\u68c0\u7d22")
        standard.append([code, label, policy, "\u4fdd\u7559\u6570\u636e\u5e93\u8bb0\u5f55\uff0c\u4f46\u4e0d\u8ba9\u7ed3\u6784\u9879\u5360\u7528\u83dc\u54c1\u641c\u7d22\u7ed3\u679c", "", ""])
    style_sheet(standard, [18, 18, 32, 28, 20, 62])
    for row in standard.iter_rows(min_row=6, max_row=6 + len(TYPE_LABELS)):
        for cell in row:
            cell.fill = section_fill

    headers = ["\u83dc\u54c1ID", "\u83dc\u54c1\u540d\u79f0", "\u98df\u5802", "\u9910\u5385/\u697c\u5c42", "\u6863\u53e3", "\u4ef7\u683c", "\u539f\u59cb\u7c7b\u578b", "\u539f\u59cb\u5206\u7c7b", "\u6700\u7ec8\u4e00\u7ea7\u4e1a\u52a1\u5206\u533a", "\u6700\u7ec8\u4e8c\u7ea7\u83dc\u54c1\u5206\u7c7b", "\u9910\u6b21\u5c5e\u6027", "\u5efa\u8bae\u5b66\u751f\u7aef\u68c0\u7d22", "\u5206\u7c7b\u8bf4\u660e", "\u6765\u6e90", "\u4eba\u5de5\u72b6\u6001"]
    data_ws.append(headers)
    for r in records:
        explanation = "\u7ed3\u6784\u9879\uff0c\u4fdd\u7559\u6e90\u8bb0\u5f55\u4f46\u4e0d\u4f5c\u83dc\u54c1" if r["item_type"] == "section" else ("\u4e0e\u4e1a\u52a1\u5206\u533a\u5e76\u5217\uff0c\u9910\u6b21\u4ec5\u4f5c\u5f85\u6838\u9a8c\u5c5e\u6027" if r["item_type"] == "meal" else "\u4f9d\u636e\u6863\u6848\u7c7b\u578b\u5f52\u7c7b\uff0c\u4e0d\u6539\u5199\u539f\u59cb\u540d\u79f0")
        data_ws.append([r["id"], r["name"], r["canteen"], r["restaurant"], r["stall"], r["price"], r["item_type"], r["original_category"], TYPE_LABELS[r["item_type"]], r["category"], r["period"], r["retrieval"], explanation, r["source"], r["review"]])
    style_sheet(data_ws, [25, 30, 20, 22, 28, 18, 14, 18, 18, 24, 16, 18, 45, 24, 30])

    review_ws.append(["\u9879\u76ee", "\u6570\u91cf", "\u5f53\u524d\u5904\u7406", "\u961f\u53cb\u4eba\u5de5\u64cd\u4f5c", "\u7981\u6b62\u4e8b\u9879"])
    section_records = [r for r in records if r["item_type"] == "section"]
    review_ws.append(["\u76ee\u5f55\u5206\u7ec4\u8bb0\u5f55", len(section_records), "\u5df2\u6539\u4e3a section\uff0c\u4e0d\u8fdb\u5b66\u751f\u68c0\u7d22", "\u786e\u8ba4\u662f\u6807\u9898/\u5206\u7ec4\uff0c\u4e0d\u606f\u8fdb\u884c\u83dc\u54c1\u5206\u7c7b", "\u4e0d\u5347\u7ea7\u4e3a meal\uff0c\u4e0d\u5199\u5165 RAG \u83dc\u54c1\u5019\u9009"])
    period_counts = Counter(r["period"] for r in records if r["item_type"] != "section")
    for period, count in sorted(period_counts.items()):
        review_ws.append([f"\u9910\u6b21\u6807\u7b7e\uff1a{period}", count, "\u4fdd\u7559\u4e3a\u5c5e\u6027\uff0c\u6807\u8bb0\u4e3a\u63a8\u6d4b\u5f85\u6838\u9a8c", "\u5411\u6863\u53e3\u6216\u98df\u5802\u8d1f\u8d23\u4eba\u786e\u8ba4\u8425\u4e1a\u65f6\u6bb5", "\u4e0d\u56e0\u9910\u6b21\u5b57\u6bb5\u5ba3\u79f0\u4eca\u65e5\u4f9b\u5e94"])
    review_ws.append(["\u5206\u7c7b\u8fb9\u754c", "", "\u91cd\u590d\u540d\u79f0\u6309ID+\u4f4d\u7f6e+\u6863\u53e3+\u4ef7\u683c\u533a\u5206", "\u62bd\u67e5\u4f4e\u4ef7\u4e38\u5b50\u3001\u9e21\u86cb\u3001\u996e\u54c1\u3001\u6c49\u5821\u5957\u9910", "\u4e0d\u4fee\u6539\u8425\u9500\u83dc\u540d\uff0c\u4e0d\u628a\u4ef7\u683c\u89c4\u683c\u5f53\u83dc\u54c1"])
    style_sheet(review_ws, [34, 12, 44, 54, 55])

    summary_ws.append(["\u7edf\u8ba1\u7ef4\u5ea6", "\u4ee3\u7801/\u5206\u7c7b", "\u6570\u91cf", "\u5360\u5168\u90e8\u8bb0\u5f55", "\u5907\u6ce8"])
    type_counts = Counter(r["item_type"] for r in records)
    cat_counts = Counter((r["item_type"], r["category"]) for r in records)
    for code in TYPE_LABELS:
        count = type_counts[code]
        summary_ws.append(["\u4e00\u7ea7\u4e1a\u52a1\u5206\u533a", f"{code} / {TYPE_LABELS[code]}", count, count / len(records), "\u5b66\u751f\u7aef\u53ea\u4ece meal \u8d77\u6b65\uff0c\u5c0f\u5403\u548c\u996e\u54c1\u7528\u660e\u786e\u5207\u6362"])
    summary_ws.append([])
    for code, categories in FINAL_CATEGORIES.items():
        for category in categories:
            summary_ws.append([f"{TYPE_LABELS[code]}\u4e8c\u7ea7\u5206\u7c7b", category, cat_counts[(code, category)], cat_counts[(code, category)] / len(records), "\u540c\u540d\u83dc\u54c1\u4e0d\u5408\u5e76\uff0c\u6309\u4f4d\u7f6e\u548c\u6863\u53e3\u4fdd\u7559"])
    summary_ws.append([])
    summary_ws.append(["\u603b\u8bb0\u5f55", "\u5168\u90e8\u76ee\u5f55\u8bb0\u5f55", len(records), 1, "\u5305\u542b 11 \u6761\u4ece\u771f\u5b9e\u5feb\u7167\u8865\u56de\u7684 section"])
    style_sheet(summary_ws, [24, 32, 12, 16, 55])
    for cell in summary_ws[1]: cell.fill = header_fill
    for row in summary_ws.iter_rows(min_row=2, min_col=4, max_col=4): row[0].number_format = "0.0%"

    handoff_ws.append(["\u89c4\u5219\u7f16\u53f7", "\u5206\u7c7b\u7ef4\u5ea6", "\u961f\u53cb\u6267\u884c\u6807\u51c6", "\u793a\u4f8b/\u5907\u6ce8"])
    rules = [
        ("R01", "\u6570\u636e\u8303\u56f4", "\u4ee5 v2 \u4e3a\u5b8c\u6574\u5e95\u7a3f\uff0c\u4e0d\u7528 modified \u8986\u76d6\u5168\u90e8\u6570\u636e", "modified \u4ec5\u4f5c\u65e7\u7248\u9910\u98df\u5f52\u7c7b\u53c2\u8003"),
        ("R02", "\u4e00\u7ea7\u5206\u533a", "\u9910\u98df\u3001\u5c0f\u5403\u3001\u996e\u54c1\u662f\u5b66\u751f\u7aef\u53ef\u5207\u6362\u7684\u516c\u5f00\u5206\u533a", "\u52a0\u8d2d\u3001\u89c4\u683c\u3001\u8d39\u7528\u3001\u76ee\u5f55\u5206\u7ec4\u6c38\u4e0d\u4f5c\u72ec\u7acb\u83dc\u54c1"),
        ("R03", "\u4e8c\u7ea7\u5206\u7c7b", "\u6309\u83dc\u54c1\u7684\u4e3b\u8981\u5f62\u6001\u548c\u70f9\u996a\u4e1a\u52a1\u5f52\u7c7b\uff0c\u4e0d\u628a\u9910\u6b21\u6df7\u8fdb\u6765", "\u9762\u98df/\u7c89\u7c7b\u3001\u7c73\u996d/\u76d6\u996d\u3001\u706b\u9505/\u9ebb\u8fa3\u70eb\u7b49"),
        ("R04", "\u9910\u6b21", "\u65e9\u9910\u3001\u5348\u9910/\u665a\u9910\u3001\u5168\u5929\u662f\u5e76\u5217\u5c5e\u6027\uff0c\u7edf\u4e00\u6807\u6ce8\u63a8\u6d4b\u5f85\u6838\u9a8c", "\u4e0d\u4f5c\u4e3a\u7b2c\u4e09\u5c42\u83dc\u54c1\u7c7b\u522b"),
        ("R05", "\u91cd\u590d\u540d\u79f0", "\u540c\u540d\u4e0d\u5408\u5e76\uff0c\u4f7f\u7528\u7a33\u5b9a ID \u548c\u5b8c\u6574\u4f4d\u7f6e\u533a\u5206", "\u4e0d\u540c\u6863\u53e3\u7684\u4e38\u5b50\u53ef\u540c\u540d\u5b58\u5728"),
        ("R06", "\u4ef7\u683c\u4e0e\u89c4\u683c", "\u4ef7\u683c\u89c4\u683c\u3001\u52a0\u6599\u548c\u8d39\u7528\u4e0d\u5347\u7ea7\u4e3a\u72ec\u7acb\u9910\u98df", "\u4e09\u4eba\u4efd\u3001\u4e38\u5b50\u3001\u52a0\u9e21\u86cb\u5e94\u5c5e\u5c5e\u6027\u6216\u52a0\u8d2d"),
        ("R07", "\u4eba\u5de5\u5ba1\u6838", "\u53ea\u4fee\u6539\u6700\u7ec8\u4e00\u7ea7\u3001\u4e8c\u7ea7\u548c\u9910\u6b21\u5b57\u6bb5\uff0c\u4e0d\u6539\u83dc\u54c1\u539f\u540d\u548c\u4ef7\u683c", "\u6709\u4e89\u8bae\u7684\u8bb0\u5f55\u8fdb\u5f85\u6838\u9a8c\uff0c\u4e0d\u9759\u9ed8\u4e22\u5f03"),
    ]
    for row in rules: handoff_ws.append(row)
    style_sheet(handoff_ws, [14, 20, 70, 50])

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)

    check = load_workbook(OUTPUT_PATH, read_only=True, data_only=True)
    assert check.sheetnames == ["\u5206\u7c7b\u6807\u51c6", "\u6700\u7ec8\u5206\u7c7b\u6570\u636e", "\u5f85\u4eba\u5de5\u6838\u9a8c", "\u7edf\u8ba1\u6458\u8981", "\u4ea4\u63a5\u89c4\u5219"]
    final_rows = list(check.worksheets[1].iter_rows(min_row=2, values_only=True))
    assert len(final_rows) == 2563
    counts = Counter(row[6] for row in final_rows)
    assert counts == Counter({"meal": 2089, "snack": 149, "beverage": 80, "addon": 221, "variant": 9, "fee": 4, "section": 11}), counts
    assert sum(1 for row in final_rows if row[11] == "\u5426") == 245
    allowed_categories = {category for categories in FINAL_CATEGORIES.values() for category in categories}
    assert all(row[9] in allowed_categories for row in final_rows)
    print(OUTPUT_PATH)
    print("rows", len(final_rows))
    print("type_counts", dict(counts))
    print("sheets", check.sheetnames)


if __name__ == "__main__":
    main()
