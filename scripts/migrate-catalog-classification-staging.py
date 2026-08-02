import argparse
import json
import shutil
import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = ROOT / "data" / "real-catalog-campus-2026-07-27-v2.sqlite"
DEFAULT_WORKBOOK = ROOT / "artifacts" / "\u83dc\u54c1\u5206\u7c7b\u6807\u51c6_\u6700\u7ec8\u7248_2026-08.xlsx"
DEFAULT_OUTPUT = ROOT / "data" / "real-catalog-classified-2026-08-02.sqlite"
PUBLIC_TYPES = {"meal", "snack", "beverage"}


def parse_args():
    parser = argparse.ArgumentParser(description="Apply audited catalog classification to an isolated SQLite copy.")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--force", action="store_true")
    return parser.parse_args()


def add_column(db, name, definition):
    existing = {row[1] for row in db.execute("PRAGMA table_info(dishes)")}
    if name not in existing:
        db.execute(f"ALTER TABLE dishes ADD COLUMN {name} {definition}")


def main():
    args = parse_args()
    for path in (args.source, args.workbook):
        if not path.exists():
            raise SystemExit(f"Missing input: {path}")
    if args.output.exists() and not args.force:
        raise SystemExit(f"Output already exists; use --force only to replace it: {args.output}")

    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_dir = ROOT / "artifacts" / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup = backup_dir / f"{args.source.stem}.before-classification-{stamp}.sqlite"
    shutil.copy2(args.source, backup)

    if args.output.exists():
        args.output.unlink()
    args.output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(args.source, args.output)

    workbook = load_workbook(args.workbook, read_only=True, data_only=True)
    rows = list(workbook["\u6700\u7ec8\u5206\u7c7b\u6570\u636e"].iter_rows(min_row=2, values_only=True))
    classifications = {row[0]: row for row in rows if row[0]}
    if len(classifications) != 2563:
        raise SystemExit(f"Expected 2563 unique workbook rows, received {len(classifications)}")

    db = sqlite3.connect(args.output)
    db.execute("PRAGMA foreign_keys = ON")
    db.execute("PRAGMA busy_timeout = 5000")
    try:
        source_ids = {row[0] for row in db.execute("SELECT id FROM dishes")}
        if source_ids != set(classifications):
            raise SystemExit("Workbook IDs do not exactly match the source database IDs")

        add_column(db, "catalog_item_type", "TEXT NOT NULL DEFAULT 'meal'")
        add_column(db, "catalog_category", "TEXT NOT NULL DEFAULT '\u5176\u4ed6\u9910\u98df'")
        add_column(db, "meal_period", "TEXT NOT NULL DEFAULT '\u5f85\u6838\u9a8c'")
        add_column(db, "review_status", "TEXT NOT NULL DEFAULT 'approved'")
        add_column(db, "retrieval_eligible", "INTEGER NOT NULL DEFAULT 1")
        db.execute("CREATE INDEX IF NOT EXISTS idx_dishes_catalog_partition ON dishes(tenant_id, catalog_item_type, catalog_category, retrieval_eligible, status)")

        db.execute("BEGIN IMMEDIATE")
        for row in classifications.values():
            dish_id, _name, _canteen, _restaurant, _stall, _price, item_type, _original, _top_label, category, period, _retrieval_label, _explanation, _source, _review = row
            is_public = item_type in PUBLIC_TYPES
            is_section = item_type == "section"
            db.execute(
                """
                UPDATE dishes
                SET catalog_item_type = ?, catalog_category = ?, meal_period = ?,
                    review_status = ?, retrieval_eligible = ?,
                    status = CASE WHEN ? THEN 'hidden' ELSE status END,
                    reservation_enabled = CASE WHEN ? THEN 0 ELSE reservation_enabled END,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (item_type, category, period, "excluded" if is_section else "approved", 1 if is_public else 0, is_section, is_section, dish_id),
            )
        db.commit()

        type_counts = dict(db.execute("SELECT catalog_item_type, COUNT(*) FROM dishes GROUP BY catalog_item_type").fetchall())
        non_public = db.execute("SELECT COUNT(*) FROM dishes WHERE retrieval_eligible = 0").fetchone()[0]
        section_count = db.execute("SELECT COUNT(*) FROM dishes WHERE catalog_item_type = 'section' AND review_status = 'excluded' AND retrieval_eligible = 0").fetchone()[0]
        if sum(type_counts.values()) != 2563 or non_public != 245 or section_count != 11:
            raise SystemExit(f"Post-migration validation failed: types={type_counts}, non_public={non_public}, sections={section_count}")
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()

    report = {
        "source": str(args.source),
        "backup": str(backup),
        "output": str(args.output),
        "workbook": str(args.workbook),
        "rows": 2563,
        "typeCounts": type_counts,
        "nonPublicCount": non_public,
        "sectionExcludedCount": section_count,
        "migratedAt": datetime.now().isoformat(timespec="seconds"),
    }
    report_path = ROOT / "artifacts" / f"catalog-classification-migration-{stamp}.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False))


if __name__ == "__main__":
    main()
